# CLI Interface for Study Planner with Multi-Format Save/Load Support and GUI Readiness

import importlib.util
import sys
import os
import json
import csv
from cryptography.fernet import Fernet
from openpyxl import Workbook, load_workbook

# Specify the full path to the core logic file
core_logic_path = r"C:\Users\nicol\CORELOGICPY.py"

try:
    if not os.path.exists(core_logic_path):
        raise FileNotFoundError(f"Core logic file not found at: {core_logic_path}")

    spec = importlib.util.spec_from_file_location("core_logic", core_logic_path)
    core = importlib.util.module_from_spec(spec)
    sys.modules["core_logic"] = core
    spec.loader.exec_module(core)

    # Import functions from the dynamically loaded module
    generate_schedule = core.generate_schedule
except FileNotFoundError as e:
    print(f"Warning: {e}. Limited functionality will be available.")
    generate_schedule = None


def get_encryption_key():
    key_path = "encryption_key.key"
    if not os.path.exists(key_path):
        key = Fernet.generate_key()
        with open(key_path, 'wb') as key_file:
            key_file.write(key)
    else:
        with open(key_path, 'rb') as key_file:
            key = key_file.read()
    return key

def encrypt_data(data, key):
    fernet = Fernet(key)
    encrypted_data = fernet.encrypt(json.dumps(data).encode())
    return encrypted_data

def decrypt_data(encrypted_data, key):
    fernet = Fernet(key)
    decrypted_data = json.loads(fernet.decrypt(encrypted_data).decode())
    return decrypted_data

def list_available_schedules():
    files = [f for f in os.listdir('.') if f.endswith(('.json', '.xlsx', '.csv'))]
    if files:
        print("\nAvailable schedules:")
        for i, file in enumerate(files, 1):
            print(f"  {i}. {file}")
    else:
        print("No saved schedules found.")
    return files

def save_schedule(schedule, filename, format="json"):
    if format == "json":
        key = get_encryption_key()
        if not filename.endswith('.json'):
            filename += '.json'
        if os.path.exists(filename):
            confirm = input(f"File '{filename}' already exists. Overwrite? (y/n): ").strip().lower()
            if confirm != 'y':
                print("Save operation cancelled.")
                return
        encrypted_data = encrypt_data(schedule, key)
        with open(filename, 'wb') as file:
            file.write(encrypted_data)
        print("Schedule saved securely in JSON format.")
    elif format == "xlsx":
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Study Schedule"
        sheet.append(["Day", "Subject", "Hours"])
        for day_schedule in schedule:
            for subject_info in day_schedule['Subjects']:
                sheet.append([day_schedule['Day'], subject_info['Subject'], subject_info['Hours']])
        workbook.save(filename)
        print("Schedule saved in Excel format.")
    elif format == "csv":
        if not filename.endswith('.csv'):
            filename += '.csv'
        with open(filename, 'w', newline='') as file:
            writer = csv.writer(file)
            writer.writerow(["Day", "Subject", "Hours"])
            for day_schedule in schedule:
                for subject_info in day_schedule['Subjects']:
                    writer.writerow([day_schedule['Day'], subject_info['Subject'], subject_info['Hours']])
        print("Schedule saved in CSV format.")

def load_schedule(file_path):
    if not os.path.exists(file_path):
        print(f"Error: File '{file_path}' not found.")
        return None
    if file_path.endswith('.json'):
        key = get_encryption_key()
        with open(file_path, 'rb') as file:
            encrypted_data = file.read()
        schedule = decrypt_data(encrypted_data, key)
        return schedule
    elif file_path.endswith('.xlsx'):
        workbook = load_workbook(file_path)
        sheet = workbook.active
        schedule = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            day, subject, hours = row
            while len(schedule) < day:
                schedule.append({"Day": len(schedule) + 1, "Subjects": []})
            schedule[day - 1]["Subjects"].append({"Subject": subject, "Hours": hours})
        return schedule
    elif file_path.endswith('.csv'):
        schedule = []
        with open(file_path, 'r') as file:
            reader = csv.DictReader(file)
            for row in reader:
                day = int(row['Day'])
                subject = row['Subject']
                hours = float(row['Hours'])
                while len(schedule) < day:
                    schedule.append({"Day": len(schedule) + 1, "Subjects": []})
                schedule[day - 1]["Subjects"].append({"Subject": subject, "Hours": hours})
        return schedule

def delete_schedule():
    files = list_available_schedules()
    if not files:
        return
    while True:
        try:
            choice = int(input("Select a schedule to delete (by number): "))
            if choice < 1 or choice > len(files):
                print("Invalid choice. Please try again.")
                continue
            filename = files[choice - 1]
            os.remove(filename)
            print(f"Schedule '{filename}' deleted.")
            break
        except (ValueError, IndexError):
            print("Invalid input. Please enter a valid number.")
        except Exception as e:
            print(f"Error: Unable to delete the schedule. {e}")

def get_user_input():
    try:
        subjects = input("Enter subjects to study (comma-separated): ").split(',')
        subjects = [subject.strip() for subject in subjects]
        total_hours = float(input("Enter total study hours available per day: "))
        num_days = int(input("Enter the number of days to plan for: "))
        if not subjects or total_hours <= 0 or num_days <= 0:
            raise ValueError("Invalid input: Ensure subjects are provided and hours/days are greater than 0.")
        return subjects, total_hours, num_days
    except ValueError as e:
        print(f"Error: {e}")
        return get_user_input()

def display_schedule(schedule):
    print("\nGenerated Study Schedule:\n")
    for day_schedule in schedule:
        print(f"Day {day_schedule['Day']}: ")
        for subject in day_schedule['Subjects']:
            print(f"  - {subject['Subject']}: {subject['Hours']} hours")
        print()

def main():
    schedule = None
    while True:
        command = input("\nEnter a command (start/save/load/delete/help/exit): ").strip().lower()
        if command == "start":
            if generate_schedule is None:
                print("Error: Cannot start. Core logic file is missing.")
                continue
            subjects, total_hours, num_days = get_user_input()
            schedule = generate_schedule(subjects, total_hours, num_days)
            display_schedule(schedule)
        elif command == "save":
            if schedule:
                format = input("Enter format to save schedule (json/xlsx/csv): ").strip().lower()
                if format not in ["json", "xlsx", "csv"]:
                    print("Invalid format. Please choose json, xlsx, or csv.")
                    continue
                filename = input("Enter filename to save schedule: ").strip()
                save_schedule(schedule, filename, format)
            else:
                print("No schedule to save. Please generate one first.")
        elif command == "load":
            files = list_available_schedules()
            if not files:
                print("No saved schedules found.")
                continue
            try:
                choice = int(input("Select a schedule to load (by number): "))
                if choice < 1 or choice > len(files):
                    print("Invalid choice. Please try again.")
                    continue
                filename = files[choice - 1]
                schedule = load_schedule(filename)
                if schedule:
                    display_schedule(schedule)
                else:
                    print("Failed to load the schedule.")
            except (ValueError, IndexError):
                print("Invalid input. Please enter a valid number.")
            except Exception as e:
                print(f"Error: Unable to load the schedule. {e}")
        elif command == "delete":
            delete_schedule()
        elif command == "help":
            print("\nCommands:")
            print("  start  - Start the study planner")
            print("  save   - Save the current schedule")
            print("  load   - Load a saved schedule")
            print("  delete - Delete the saved schedule")
            print("  help   - Display this help message")
            print("  exit   - Exit the program")
        elif command == "exit":
            print("Exiting program. Goodbye!")
            break
        else:
            print("Unknown command. Type 'help' for a list of available commands.")

if __name__ == "__main__":
      main()
